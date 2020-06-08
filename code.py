import pickle
from openpyxl import load_workbook
import pprint

wb = load_workbook('KIKmix.20191001(말소코드포함).xlsx')
ws = wb['KIKmix']

code = []
temp = []
for x in ws['E']:
    code.append(x.value)

for x in code:
    if x != None:
        temp.append(x)


with open('code.txt', 'wb') as fp:
    pickle.dump(temp, fp)

with open('code.txt', 'rb') as fp:
    arc = pickle.load(fp)

pprint.pprint(arc)

#code that downloads codes from excel
